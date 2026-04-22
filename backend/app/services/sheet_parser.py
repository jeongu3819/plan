"""
Excel/CSV → Sheet 구조 파싱 서비스
- 병합셀 구조 유지
- 색상 정보 추출
- 체크 가능 셀 자동 감지 (O/X/빈칸/체크마크 패턴)
- 수식은 결과값 기준 표시
- 줄바꿈/섹션/제목 유지
- v3.1: 컬럼 역할 자동 추정 (체크 여부/점검일시/담당자/예정일/비고)
"""

import io
import csv
import hashlib
import logging
import re
from datetime import datetime
from typing import Tuple, Optional, Dict, List, Any

logger = logging.getLogger("sheet_parser")

# 체크 가능 "값" 패턴 (빈 셀 제외 — 빈 셀만으로는 체크 의도를 확정할 수 없음)
CHECKABLE_VALUES = {"o", "O", "x", "X", "○", "●", "×", "✓", "✔", "☑", "☐", "□", "■", "△", "▲", "-"}

# 컬럼 헤더에서 "체크 컬럼"을 식별할 키워드
CHECK_HEADER_KEYWORDS = ("체크", "점검결과", "판정", "결과", "상태", "확인", "check", "status", "result", "ok/ng", "pass/fail")

# 컬럼 헤더에서 "라벨(점검항목) 컬럼"을 식별할 키워드
LABEL_HEADER_KEYWORDS = ("점검항목", "항목", "내용", "설명", "검사항목", "item", "description", "task")

# 실제 체크 값으로 많이 쓰이는 완료/미완료 계열 (한국어 체크시트 관행)
CHECK_STATE_VALUES = {"완료", "미완료", "진행", "진행 중", "진행중", "보류", "불가", "ok", "OK", "ng", "NG", "n/a", "N/A", "해당없음", "pass", "PASS", "fail", "FAIL", "양호", "불량"}

# ═══════════════════════════════════════════════════════════
# v3.1  컬럼 역할 자동 추정 엔진
# ═══════════════════════════════════════════════════════════

COLUMN_ROLE_HEADER_CANDIDATES: Dict[str, List[str]] = {
    "check_status": [
        # 직접적 체크/점검 여부
        "체크 여부", "체크여부", "체크 유/무", "체크유무", "체크 유무",
        "점검 여부", "점검여부", "확인 여부", "확인여부",
        # 진행/수행/완료 여부
        "진행 여부", "진행여부", "진행 유/무", "진행유무", "진행 유무",
        "수행 여부", "수행여부", "완료 여부", "완료여부", "완료", "수행",
        # 상태/결과/판정
        "상태", "결과", "점검결과", "판정", "점검 결과",
        "Check", "Checked", "Status", "Result", "OK/NG", "Pass/Fail",
    ],
    "checked_at": [
        "실제 점검일시", "실제점검일시", "점검일시", "점검일", "점검일자",
        "확인일시", "확인일", "완료일시", "완료일", "작업일시", "작업일",
        "실제점검일", "수행일시", "수행일",
        "Checked At", "Completed At", "Inspected At", "Done At",
    ],
    "progress_date": [
        # 진행일 전용 role — checked_at과 분리해서 상태-진행일 짝으로 다루기 위함
        "진행일", "진행 일자", "진행일자", "진행 일시", "진행일시",
        "실시일", "실시일자", "이행일", "이행일자",
        "Progress Date", "Done Date",
    ],
    "assignee": [
        "담당자", "근무자", "작업자", "점검자", "확인자", "수행자", "실시자",
        "책임자", "관리자",
        "Operator", "Inspector", "Assignee", "Checker", "Owner",
    ],
    "due_date": [
        "점검예정일", "예정일", "점검일정", "계획일", "목표일",
        "예정일시", "예정일자", "계획일자", "스케줄",
        "Due Date", "Scheduled", "Planned",
    ],
    "remark": [
        "비고", "메모", "특이사항", "참고사항", "이상내용", "조치사항", "조치내용",
        "기타", "참고", "코멘트", "의견",
        "Note", "Remark", "Comment", "Memo",
    ],
}

# 값 패턴 (보조 판단)
_DATE_RE = re.compile(
    r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}"  # 2026-04-19 ...
    r"|^\d{1,2}[-/.]\d{1,2}[-/.]\d{4}"  # 19-04-2026 ...
    r"|^\d{1,2}[-/.]\d{1,2}"            # 04-19
)
_CHECK_VALUE_SET = {"O", "o", "X", "x", "○", "●", "×", "✓", "✔", "☑", "☐",
                    "완료", "미완료", "Y", "N", "TRUE", "FALSE", "정상", "이상", "보류",
                    "양호", "불량", "OK", "NG", "PASS", "FAIL"}


def _header_score(header_text: str, candidates: List[str]) -> float:
    """헤더명과 후보 목록 사이의 유사도 점수(0~1)"""
    if not header_text:
        return 0.0
    h = header_text.strip()
    h_lower = h.lower()
    for cand in candidates:
        cl = cand.lower()
        if h_lower == cl:
            return 1.0
        if cl in h_lower or h_lower in cl:
            return 0.8
    # 부분 키워드 매칭
    for cand in candidates:
        tokens = cand.lower().split()
        if any(t in h_lower for t in tokens if len(t) >= 2):
            return 0.5
    return 0.0


def _value_pattern_score(values: List[str], role: str) -> float:
    """컬럼의 실제 값 패턴을 분석하여 역할 적합 점수(0~1)"""
    if not values:
        return 0.0
    non_empty = [v for v in values if v.strip()]
    total = len(values)
    filled = len(non_empty)

    if role == "check_status":
        check_count = sum(1 for v in non_empty if v.strip() in _CHECK_VALUE_SET)
        empty_ratio = 1 - (filled / total) if total > 0 else 0
        if filled > 0 and check_count / filled >= 0.5:
            return 0.8
        if empty_ratio >= 0.3 and filled > 0 and check_count >= 1:
            return 0.5
        return 0.0

    elif role == "checked_at":
        date_count = sum(1 for v in non_empty if _DATE_RE.match(v.strip()))
        empty_ratio = 1 - (filled / total) if total > 0 else 0
        if filled > 0 and date_count / filled >= 0.5:
            return 0.8
        if empty_ratio >= 0.3 and date_count >= 1:
            return 0.6
        return 0.0

    elif role == "assignee":
        # 담당자: 보통 짧은 한글 이름 2~4자, 대부분 채워져 있음
        name_like = sum(1 for v in non_empty if 1 < len(v.strip()) <= 10 and not v.strip().isdigit())
        if filled > 0 and name_like / filled >= 0.5:
            return 0.6
        return 0.0

    elif role == "due_date":
        date_count = sum(1 for v in non_empty if _DATE_RE.match(v.strip()))
        if filled > 0 and date_count / filled >= 0.5:
            return 0.7
        return 0.0

    elif role == "progress_date":
        # 진행일: 날짜 또는 비어있는 셀이 많음 (아직 진행 안 된 행이 있을 수 있으니까)
        date_count = sum(1 for v in non_empty if _DATE_RE.match(v.strip()))
        empty_ratio = 1 - (filled / total) if total > 0 else 0
        if filled > 0 and date_count / filled >= 0.5:
            return 0.8
        if empty_ratio >= 0.3 and date_count >= 1:
            return 0.5
        return 0.0

    elif role == "remark":
        # 비고: 대부분 비어있고, 채워진 건 10자 이상 텍스트
        if total > 0 and filled / total <= 0.3:
            return 0.5
        return 0.0

    return 0.0


def detect_column_roles(
    col_headers: Dict[int, str],
    ws=None,
    header_row: int = 1,
    total_rows: int = 0,
    sample_limit: int = 50,
) -> Dict[str, Any]:
    """
    컬럼 역할 자동 추정.

    Returns: {
        "check_status": {"col": 8, "header": "체크 여부", "confidence": 0.95},
        "checked_at": {"col": 6, "header": "실제 점검일시", "confidence": 0.90},
        ...
    }

    개선 사항 (v3.3):
    - 한 컬럼이 여러 역할의 후보가 되면 가장 점수가 높은 역할만 가져감 (exclusive)
    - 부분 매칭(0.5)만으로는 role 확정하지 않음 — 값 패턴 보강 필요 (0.5 임계)
    """
    # (col, role) → (h_score, v_score, combined) 표 작성
    candidate_table: List[Tuple[int, str, float, float, float, str]] = []
    for role, candidates in COLUMN_ROLE_HEADER_CANDIDATES.items():
        for col_idx, header_text in col_headers.items():
            h_score = _header_score(header_text, candidates)
            if h_score <= 0:
                continue
            v_score = 0.0
            if ws:
                sample_values = []
                scan_end = min(total_rows + 1, header_row + sample_limit + 1)
                for r in range(header_row + 1, scan_end):
                    cell = ws.cell(row=r, column=col_idx)
                    v = cell.value
                    sample_values.append(str(v).strip() if v is not None else "")
                v_score = _value_pattern_score(sample_values, role)
            combined = h_score * 0.7 + v_score * 0.3
            candidate_table.append((col_idx, role, h_score, v_score, combined, header_text))

    # 점수 내림차순 정렬
    candidate_table.sort(key=lambda x: x[4], reverse=True)

    used_cols: set = set()
    used_roles: set = set()
    roles: Dict[str, Any] = {}

    for col_idx, role, h_score, v_score, combined, header_text in candidate_table:
        if col_idx in used_cols or role in used_roles:
            continue
        # 부분 매칭(0.5 미만)일 때는 값 패턴이 뒷받침될 때만 채택
        if h_score < 0.8 and v_score < 0.4:
            # 단, 다른 역할이 이 컬럼을 이미 가져갔거나 가져갈 가능성이 있으면 보류
            continue
        # 정확/포함 매칭(>=0.8)이거나 부분매칭이지만 값 패턴이 일치하면 채택
        if combined < 0.45:
            continue
        roles[role] = {
            "col": col_idx - 1,
            "header": header_text,
            "confidence": round(combined, 2),
        }
        used_cols.add(col_idx)
        used_roles.add(role)

    return roles


def compute_structure_hash(col_headers: Dict[int, str]) -> str:
    """컬럼 헤더 구조의 해시를 계산하여 같은 양식 인식에 활용"""
    key_parts = sorted(f"{c}:{h.strip()}" for c, h in col_headers.items() if h.strip())
    raw = "|".join(key_parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def _color_to_hex(color) -> Optional[str]:
    """openpyxl Color 객체 → #RRGGBB 문자열"""
    if color is None:
        return None
    try:
        if color.type == "rgb" and color.rgb and color.rgb != "00000000":
            rgb = str(color.rgb)
            if len(rgb) == 8:
                return f"#{rgb[2:]}"
            elif len(rgb) == 6:
                return f"#{rgb}"
        if color.type == "theme":
            return None  # 테마 색상은 클라이언트에서 처리
        if color.type == "indexed":
            return None
    except Exception:
        pass
    return None


def _get_cell_bg_color(cell) -> Optional[str]:
    """셀 배경색 추출"""
    try:
        fill = cell.fill
        if fill and fill.fgColor:
            c = _color_to_hex(fill.fgColor)
            if c and c != "#000000":
                return c
        if fill and fill.bgColor:
            c = _color_to_hex(fill.bgColor)
            if c and c != "#000000":
                return c
    except Exception:
        pass
    return None


def _get_cell_font_info(cell) -> dict:
    """셀 폰트 정보 추출"""
    info = {}
    try:
        font = cell.font
        if font:
            if font.bold:
                info["bold"] = True
            if font.italic:
                info["italic"] = True
            if font.size:
                info["fontSize"] = font.size
            color = _color_to_hex(font.color)
            if color:
                info["fontColor"] = color
    except Exception:
        pass
    return info


def _get_cell_borders(cell) -> dict:
    """셀 테두리 정보"""
    borders = {}
    try:
        b = cell.border
        for side_name in ("top", "bottom", "left", "right"):
            side = getattr(b, side_name, None)
            if side and side.style:
                borders[side_name] = side.style
    except Exception:
        pass
    return borders


def _header_matches(header_text: str, keywords) -> bool:
    if not header_text:
        return False
    h = header_text.strip().lower()
    return any(k.lower() in h for k in keywords)


# ═══════════════════════════════════════════════════════════
# 상태 컬럼 값 정규화
#   "O" / "X" / "N/A" 같은 단일 토큰은 그대로 두되
#   "O(신규 bot 장착)" / "X - 사유" / "O: 설명" 같은 패턴은
#   상태와 비고로 분리한다. 일반 텍스트("완료" 같은 한글 상태어)는 보존.
# ═══════════════════════════════════════════════════════════
# 길이가 긴 토큰을 먼저 둬야 정규식 alternation이 OK > O 처럼 동작 (그렇지 않으면 "OK" 가 "O" + "K" 로 잘림)
_STATUS_PREFIX_RE = re.compile(
    r"^\s*(?P<status>"
    r"진행\s*중|진행중"
    r"|해당없음|미완료|완료|보류|양호|불량|정상|이상"
    r"|N\s*/\s*A|n\s*/\s*a|N\.A\.?|n\.a\.?"
    r"|PASS|pass|FAIL|fail|OK|ok|NG|ng"
    r"|○|●|×|△|▲"
    r"|O|o|X|x"
    r")",
    re.UNICODE,
)


def _normalize_status_value(raw) -> Tuple[str, str]:
    """
    상태(체크) 컬럼의 셀 값을 (parsed_status, parsed_note)로 분리.

    예:
      "O"                  → ("O", "")
      "X"                  → ("X", "")
      "N/A"                → ("N/A", "")
      "O(신규 bot 장착)"     → ("O", "신규 bot 장착")
      "X - 사유"            → ("X", "사유")
      "O: 설명"             → ("O", "설명")
      "완료"                → ("완료", "")
      ""                   → ("", "")
      알 수 없는 텍스트       → (원본, "")
    """
    if raw is None:
        return ("", "")
    s = str(raw).strip()
    if not s:
        return ("", "")

    m = _STATUS_PREFIX_RE.match(s)
    if not m:
        return (s, "")

    status_token = m.group("status")
    rest = s[m.end():]

    # status_token 뒤에 공백/구분자/괄호 없이 곧바로 다른 글자가 이어지면(예: "OK점검완")
    # 의도적으로 분리하지 않는다 — 원문 보존.
    if rest and rest[0].isalnum() and not status_token.isalpha():
        # status가 "O" 같은 단일 글자인데 그 뒤가 alphanumeric이면 보통 그냥 토큰의 일부 (드문 케이스)
        # 안전하게 원본 유지
        return (s, "")

    # 상태 표준화
    su = status_token.upper().replace(" ", "").replace(".", "")
    if status_token == "○" or su == "O":
        status = "O"
    elif status_token == "●":
        status = "O"
    elif status_token == "×" or su == "X":
        status = "X"
    elif su in ("NA", "N/A"):
        status = "N/A"
    elif status_token == "해당없음":
        status = "N/A"
    elif status_token in ("△", "▲"):
        status = status_token
    else:
        # 완료/미완료/OK/NG/양호 등은 입력 그대로 유지
        status = status_token

    note = ""
    if rest:
        rest_str = rest.strip()
        # 괄호로 감싼 비고
        paren_match = re.match(r"^[\(\[（【]\s*(.*?)\s*[\)\]）】]\s*$", rest_str)
        if paren_match:
            note = paren_match.group(1).strip()
        else:
            # 선행 구분자(-, :, –, —, /, 공백) 제거
            note = re.sub(r"^[\s\-:：–—/]+", "", rest_str).strip()

    return (status, note)


def _is_checkable_cell(value, col_header: str = "") -> bool:
    """
    셀이 체크 가능한 항목인지 판단.
    - 컬럼 헤더가 체크 키워드 포함 → 비어 있지 않은 셀 전부 체크 대상 (완료/미완료/OK/NG 등 포함)
    - 그렇지 않으면 값 자체가 체크 마크 문자거나 상태 키워드여야 함
    """
    if value is None:
        v = ""
    else:
        v = str(value).strip()

    if _header_matches(col_header, CHECK_HEADER_KEYWORDS):
        return v != "" or True  # 헤더가 명시적 체크 컬럼이면 행 자체가 체크 항목
    if not v:
        return False
    if v in CHECKABLE_VALUES:
        return True
    if v in CHECK_STATE_VALUES:
        return True
    return False


def _find_label_for_cell(ws, row_idx: int, col_idx: int, label_cols=None) -> str:
    """
    체크 셀에 대한 라벨 찾기.
    - label_cols 가 주어지면 그 컬럼들에서 같은 행 값을 우선 사용
    - 없으면 같은 행의 왼쪽 텍스트 셀을 스캔
    """
    if label_cols:
        for lc in label_cols:
            cell = ws.cell(row=row_idx, column=lc)
            v = cell.value
            if v is not None and str(v).strip():
                return str(v).strip()[:200]
    for c in range(col_idx - 1, 0, -1):
        cell = ws.cell(row=row_idx, column=c)
        v = cell.value
        if v is not None and str(v).strip():
            return str(v).strip()[:200]
    return ""


# 헤더로 자주 등장하는 키워드 (전체 컬럼 후보 합집합 + 일반 키워드)
_HEADER_KEYWORD_BONUS = {
    "분류", "항목", "내용", "설명", "구분", "번호", "no", "no.",
    "담당", "담당자", "근무자", "작업자", "점검자", "확인자", "수행자",
    "상태", "결과", "여부", "유무", "체크", "점검", "확인", "진행", "수행",
    "일자", "일시", "날짜", "예정", "이행", "진행일", "점검일",
    "비고", "메모", "특이", "참고", "코멘트",
    "약품", "설비", "장비", "부품", "코드", "이름",
    "check", "item", "status", "result", "name", "code", "date", "note", "remark",
}

# 헤더로 인식되면 안 되는 데이터값 (정규화 후 비교)
_DATA_VALUE_TOKENS = {
    "o", "x", "○", "●", "×", "✓", "✔", "☑", "☐", "□", "■", "△", "▲",
    "n/a", "na", "ok", "ng", "pass", "fail", "y", "n", "true", "false",
    "완료", "미완료", "진행", "보류", "양호", "불량", "정상", "이상", "해당없음",
    "-",
}


def _is_pure_data_token(text: str) -> bool:
    """`O`, `X`, `N/A` 같은 짧은 데이터값 토큰인지 판정. 헤더면 항상 False여야 함."""
    if not text:
        return False
    t = text.strip().lower().replace(" ", "")
    if t in _DATA_VALUE_TOKENS:
        return True
    # 한 글자 영문/숫자/기호만 있어도 데이터값으로 본다
    if len(t) <= 1:
        return True
    return False


def _row_header_score(ws, row: int, total_cols: int) -> Tuple[float, int, int]:
    """
    한 행이 '헤더 행'일 가능성 점수.
    Returns (score, text_count, data_token_count)
      - score: 비어있지 않은 텍스트 셀 수 + 헤더 키워드 보너스 - 데이터값 패널티
      - text_count: 비어있지 않은 텍스트 셀 수 (가중 전)
      - data_token_count: O/X/N/A 같은 데이터값 토큰 수
    """
    text_count = 0
    keyword_hits = 0
    data_hits = 0
    long_text_count = 0
    numeric_count = 0
    for c in range(1, total_cols + 1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        if isinstance(v, (int, float)):
            numeric_count += 1
            continue
        if not isinstance(v, str):
            continue
        s = v.strip()
        if not s:
            continue
        text_count += 1
        if _is_pure_data_token(s):
            data_hits += 1
            continue
        s_lower = s.lower()
        # 키워드 가산
        if any(kw in s_lower for kw in _HEADER_KEYWORD_BONUS):
            keyword_hits += 1
        # 긴 한글/영문 텍스트 (헤더보다는 데이터인 경우가 많음)
        if len(s) > 20:
            long_text_count += 1
    # 점수: 텍스트 셀 수 기본값에 키워드 보너스 큰 가중, 데이터값/긴문장은 패널티
    score = text_count + keyword_hits * 2.5 - data_hits * 1.5 - long_text_count * 0.5 - numeric_count * 0.3
    return score, text_count, data_hits


def _detect_header_row(ws, max_scan: int = 10) -> int:
    """
    헤더 행(1-based) 자동 감지:
    - 키워드 가중 점수가 가장 높은 행을 선택
    - 'O', 'X', 'N/A' 같은 데이터값 토큰이 다수면 헤더 후보에서 감점
    - 동률이면 더 아래 행 선호 (제목 행 다음에 실제 헤더가 오는 패턴)
    """
    total_cols = ws.max_column or 0
    if total_cols == 0:
        return 1
    best_row = 1
    best_score = -1e9
    scan_to = min(max_scan, ws.max_row or 0)
    for r in range(1, scan_to + 1):
        score, text_count, data_hits = _row_header_score(ws, r, total_cols)
        # 텍스트가 거의 없으면 후보에서 제외 (제목 한 줄짜리 행)
        if text_count < 2:
            continue
        # 데이터값이 텍스트의 절반 이상이면 데이터 행으로 간주 → 후보 제외
        if data_hits > 0 and data_hits >= max(2, text_count // 2):
            continue
        if score >= best_score:
            best_score = score
            best_row = r
    # 만약 모든 행이 제외됐으면 fallback: 텍스트 셀 가장 많은 행
    if best_score == -1e9:
        for r in range(1, scan_to + 1):
            count = 0
            for c in range(1, total_cols + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and isinstance(v, str) and v.strip():
                    count += 1
            if count > 0:
                return r
        return 1
    return best_row


def _build_column_headers(ws, header_row: int) -> dict:
    """header_row 의 각 컬럼 텍스트를 dict {col_idx(1-based): header_text} 로 반환"""
    headers = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip():
            headers[c] = str(v).strip()
    return headers


def parse_excel_to_structure(
    file_bytes: bytes,
    filename: str,
    target_sheet: Optional[str] = None,
) -> Tuple[dict, dict]:
    """
    Excel/CSV → (structure, meta) 반환

    structure: {
        cells: [{row, col, value, type, bg, font, borders, rowSpan, colSpan}],
        merges: [{startRow, startCol, endRow, endCol}],
        col_widths: [float, ...],
        row_heights: [float, ...],
        total_rows: int,
        total_cols: int,
        checkable_cells: [{ref, row, col, label}],
        headers: [{col, value}],
    }
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        return _parse_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        return _parse_xlsx(file_bytes, target_sheet)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: .{ext} (xlsx, csv만 지원)")


def _parse_csv(file_bytes: bytes) -> Tuple[dict, dict]:
    """CSV 파싱"""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows_data = list(reader)

    if not rows_data:
        return {"cells": [], "merges": [], "total_rows": 0, "total_cols": 0, "checkable_cells": [], "headers": []}, \
               {"row_count": 0, "col_count": 0, "checkable_count": 0, "sheet_name": None}

    max_cols = max(len(r) for r in rows_data)
    cells = []
    checkable_cells = []

    # CSV 의 첫 행을 헤더로 가정
    header_values = rows_data[0] if rows_data else []
    csv_check_cols = {
        c_idx for c_idx, h in enumerate(header_values)
        if _header_matches(h, CHECK_HEADER_KEYWORDS)
    }
    csv_label_cols = [
        c_idx for c_idx, h in enumerate(header_values)
        if _header_matches(h, LABEL_HEADER_KEYWORDS)
    ]

    for r_idx, row in enumerate(rows_data):
        for c_idx in range(max_cols):
            value = row[c_idx] if c_idx < len(row) else ""
            cells.append({
                "row": r_idx, "col": c_idx,
                "value": value, "type": "text",
            })
            if r_idx > 0:
                is_check = (c_idx in csv_check_cols) or _is_checkable_cell(value)
                if is_check:
                    label = ""
                    if csv_label_cols:
                        for lc in csv_label_cols:
                            lv = row[lc] if lc < len(row) else ""
                            if lv.strip():
                                label = lv.strip()[:200]
                                break
                    if not label:
                        for lc in range(c_idx - 1, -1, -1):
                            lv = row[lc] if lc < len(row) else ""
                            if lv.strip():
                                label = lv.strip()[:200]
                                break
                    checkable_cells.append({
                        "ref": f"{chr(65 + c_idx)}{r_idx + 1}" if c_idx < 26 else f"C{c_idx}R{r_idx + 1}",
                        "row": r_idx, "col": c_idx, "label": label,
                        "initial_value": value,
                    })

    headers = []
    if rows_data:
        for c_idx, v in enumerate(rows_data[0]):
            if v.strip():
                headers.append({"col": c_idx, "value": v.strip()})

    structure = {
        "cells": cells,
        "merges": [],
        "col_widths": [],
        "row_heights": [],
        "total_rows": len(rows_data),
        "total_cols": max_cols,
        "checkable_cells": checkable_cells,
        "headers": headers,
    }
    meta = {
        "row_count": len(rows_data),
        "col_count": max_cols,
        "checkable_count": len(checkable_cells),
        "sheet_name": None,
    }
    return structure, meta


def _parse_xlsx(file_bytes: bytes, target_sheet: Optional[str] = None) -> Tuple[dict, dict]:
    """XLSX 파싱 (openpyxl)"""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if target_sheet and target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
    else:
        ws = wb.active
    sheet_name = ws.title

    # 병합 셀 정보
    merges = []
    merged_cells_map = {}  # (row, col) -> merge info
    for mg in ws.merged_cells.ranges:
        merge_info = {
            "startRow": mg.min_row - 1,  # 0-based
            "startCol": mg.min_col - 1,
            "endRow": mg.max_row - 1,
            "endCol": mg.max_col - 1,
        }
        merges.append(merge_info)
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r == mg.min_row and c == mg.min_col:
                    merged_cells_map[(r, c)] = {
                        "rowSpan": mg.max_row - mg.min_row + 1,
                        "colSpan": mg.max_col - mg.min_col + 1,
                    }
                else:
                    merged_cells_map[(r, c)] = "hidden"

    # 열 너비
    col_widths = []
    for c in range(1, ws.max_column + 1 if ws.max_column else 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        col_widths.append(round(dim.width, 1) if dim and dim.width else 8.43)

    # 행 높이
    row_heights = []
    for r in range(1, ws.max_row + 1 if ws.max_row else 1):
        dim = ws.row_dimensions.get(r)
        row_heights.append(round(dim.height, 1) if dim and dim.height else 15.0)

    cells = []
    checkable_cells = []

    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0

    # 헤더 행 자동 감지 + 체크/라벨 컬럼 식별
    header_row = _detect_header_row(ws) if total_rows > 0 else 1
    col_headers = _build_column_headers(ws, header_row)

    # v3.2: 컬럼 역할 자동 추정을 셀 파싱 전에 수행하여
    #       "담당자/진행일/비고" 등 명시적 역할 컬럼은 체크 대상에서 제외할 수 있도록 한다.
    column_roles = detect_column_roles(col_headers, ws, header_row, total_rows)

    # v3.4: 체크 가능 컬럼은 명시적 check_status 역할로 판정된 컬럼만 사용한다.
    #   이전에는 CHECK_HEADER_KEYWORDS 부분 매칭("check 항목" 등)도 포함시켜
    #   설명 컬럼에까지 체크박스가 생성되고, 모수가 2배로 잡히는 문제가 있었다.
    if "check_status" in column_roles:
        check_cols = {column_roles["check_status"]["col"] + 1}
    else:
        check_cols = set()
    label_cols_list = [col for col, h in col_headers.items() if _header_matches(h, LABEL_HEADER_KEYWORDS)]

    for r in range(1, total_rows + 1):
        for c in range(1, total_cols + 1):
            merge_info = merged_cells_map.get((r, c))
            if merge_info == "hidden":
                continue  # 병합된 숨겨진 셀은 건너뜀

            cell = ws.cell(row=r, column=c)
            value = cell.value
            if value is not None:
                if hasattr(value, "strftime"):
                    try:
                        value = value.strftime("%Y-%m-%d %H:%M") if hasattr(value, "hour") else value.strftime("%Y-%m-%d")
                    except Exception:
                        value = str(value)
                else:
                    value = str(value)
            else:
                value = ""

            cell_data = {
                "row": r - 1,  # 0-based
                "col": c - 1,
                "value": value,
                "type": "text",
            }

            # 병합 span
            if isinstance(merge_info, dict):
                cell_data["rowSpan"] = merge_info["rowSpan"]
                cell_data["colSpan"] = merge_info["colSpan"]

            # 배경색
            bg = _get_cell_bg_color(cell)
            if bg:
                cell_data["bg"] = bg

            # 폰트
            font_info = _get_cell_font_info(cell)
            if font_info:
                cell_data["font"] = font_info

            # 테두리
            borders = _get_cell_borders(cell)
            if borders:
                cell_data["borders"] = borders

            # 정렬
            try:
                if cell.alignment:
                    if cell.alignment.horizontal:
                        cell_data["align"] = cell.alignment.horizontal
                    if cell.alignment.wrap_text:
                        cell_data["wrapText"] = True
            except Exception:
                pass

            cells.append(cell_data)

            # 체크 가능 셀 감지 (헤더 행 이후만)
            #   v3.4: check_status 역할로 명시 판정된 컬럼만 체크 대상.
            #         설명/항목 컬럼이나 값 패턴만으로 잡히는 셀은 절대 체크 대상이 아니다.
            if r > header_row and c in check_cols:
                label = _find_label_for_cell(ws, r, c, label_cols_list)
                col_letter = get_column_letter(c)
                entry = {
                    "ref": f"{col_letter}{r}",
                    "row": r - 1,
                    "col": c - 1,
                    "label": label,
                    "initial_value": value,  # 기존 "완료"/"미완료" 등 상태값 보존
                }
                if value:
                    parsed_status, parsed_note = _normalize_status_value(value)
                    if parsed_status:
                        entry["parsed_status"] = parsed_status
                    if parsed_note:
                        entry["parsed_note"] = parsed_note
                checkable_cells.append(entry)

    # 헤더 정보 (감지된 헤더 행 기준)
    headers = []
    for col, text in col_headers.items():
        headers.append({"col": col - 1, "value": text})

    # column_roles는 셀 파싱 전에 이미 계산함
    structure_hash = compute_structure_hash(col_headers)

    # ───────────────────────────────────────────────────────────
    # v3.3: 운영 기록용 가상 컬럼 자동 생성 (진행일자 → 비고 순으로 끝에 붙임)
    #   - inspection 시트에는 무조건 두 컬럼이 있어야 함 (사용자 요구)
    #   - 원본에 이미 있으면 (column_roles에 등록됨) 추가하지 않음
    #   - 진행일자: progress_date 또는 checked_at 둘 중 하나라도 있으면 OK
    #   - 비고: remark가 있으면 OK
    # ───────────────────────────────────────────────────────────
    has_status = "check_status" in column_roles

    # 1) 진행일자 가상 컬럼 — check_status가 없어도 항상 생성 (사용자가 직접 기록 가능하게)
    if "progress_date" not in column_roles and "checked_at" not in column_roles:
        virtual_col_idx = total_cols
        cells.append({
            "row": header_row - 1,
            "col": virtual_col_idx,
            "value": "진행일자",
            "type": "text",
            "font": {"bold": True, "fontColor": "#6B7280"},
            "align": "center",
        })
        headers.append({"col": virtual_col_idx, "value": "진행일자"})
        col_widths.append(14.0)
        column_roles["progress_date"] = {
            "col": virtual_col_idx,
            "header": "진행일자",
            "confidence": 1.0,
            "virtual": True,
            "editor_type": "date",
        }
        total_cols += 1

    # 2) 비고 가상 컬럼 — 항상 생성 (특이사항 기록용)
    extracted_notes = [(cc["row"], cc.get("parsed_note")) for cc in checkable_cells if cc.get("parsed_note")]
    if "remark" not in column_roles:
        virtual_col_idx = total_cols
        cells.append({
            "row": header_row - 1,
            "col": virtual_col_idx,
            "value": "비고",
            "type": "text",
            "font": {"bold": True, "fontColor": "#6B7280"},
            "align": "center",
        })
        # 상태 컬럼에서 분리된 note가 있으면 자동으로 채워줌
        for r0, note in extracted_notes:
            cells.append({
                "row": r0,
                "col": virtual_col_idx,
                "value": note,
                "type": "text",
                "font": {"fontColor": "#374151", "italic": True},
                "wrapText": True,
            })
        headers.append({"col": virtual_col_idx, "value": "비고"})
        col_widths.append(20.0)
        column_roles["remark"] = {
            "col": virtual_col_idx,
            "header": "비고",
            "confidence": 1.0,
            "virtual": True,
            "editor_type": "text",
        }
        total_cols += 1

    # has_status는 추후 metadata 표시용으로 reserved (현재 미사용 경고 회피)
    _ = has_status

    wb.close()

    # ───────────────────────────────────────────────────────────
    # v3.2: 시트 유형 자동 추천 (inspection vs assignment_mapping)
    # ───────────────────────────────────────────────────────────
    suggested_type = "inspection"
    header_texts = [h.strip() for h in col_headers.values() if h]
    assignment_keywords = ["적용설비", "설비명", "장비", "약품명", "약품", "코드", "적용 부분"]
    
    # 헤더 중 적용설비 1, 적용설비 2 식으로 나열되어 있는지 확인
    assignment_cols_count = sum(1 for h in header_texts if "설비" in h or "장비" in h or "부품" in h)
    
    if assignment_cols_count >= 2 or any("약품" in h for h in header_texts[:2]):
        suggested_type = "assignment_mapping"

    structure = {
        "cells": cells,
        "merges": merges,
        "col_widths": col_widths,
        "row_heights": row_heights,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "checkable_cells": checkable_cells,
        "headers": headers,
        "column_roles": column_roles,
        "header_row_idx": header_row - 1,   # 0-based
        "data_start_row": header_row,        # 0-based (header_row 다음 행부터 데이터)
        "structure_hash": structure_hash,
        "suggested_type": suggested_type,
    }
    meta = {
        "row_count": total_rows,
        "col_count": total_cols,
        "checkable_count": len(checkable_cells),
        "sheet_name": sheet_name,
        "structure_hash": structure_hash,
        "suggested_type": suggested_type,
    }
    return structure, meta
